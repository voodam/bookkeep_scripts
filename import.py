#!/usr/bin/env python3

import sys
import os
import re
from datetime import datetime
from decimal import Decimal
from openpyxl import load_workbook
import json
from jsoncomment import JsonComment
json_comment = JsonComment(json)
import util

_, xslx = sys.argv
account_sheet, transactions_sheet = load_workbook(filename = xslx).worksheets
with open("categories.json") as f:
  default_categories = json_comment.load(f)

def get_target(row):
  target = row[5].value
  target = re.sub("^გადახდა - ", "", target) # "Payment - "
  target = re.sub(" \d+\.\d{2} GEL \d{2}.\d{2}.\d{4}$", "", target)
  if target == "Personal Transfer.":
    beneficiary_name = row[6].value
    beneficiary_acc_num = row[7].value
    target = f"Personal transfer to {beneficiary_name} [{beneficiary_acc_num}]"
  return target

def get_defaults(target):
  account = "u"
  for regexp in default_categories.keys():
    if re.search(regexp, target, re.IGNORECASE):
      category = default_categories[regexp]
      if isinstance(category, list):
        category, account = category
      break
  else:
    category = "unknown"
  return category, account

internal_transfer_targets = ["Конвертация суммы", "Convert amount", "ანაბრის გახსნა", "ავტომატური შევსების ოპერაცია", "უნაღდო კონვერტაცია", "დეპოზიტის თანხის გატანა"]
ignore_ibans = os.environ.get("IGNORE_IBANS")
ignore_ibans = ignore_ibans.split(",") if ignore_ibans else []
skip_targets_regexp = util.seq_to_regexp(internal_transfer_targets + ignore_ibans)

iban = account_sheet.cell(row=3, column=2).value
currency = account_sheet.cell(row=4, column=2).value

data = []
iter = transactions_sheet.iter_rows()
next(iter) # skip title
for row in iter:
  if not row[2].value:
    continue # skip incomes
  target = get_target(row)

  if re.search(skip_targets_regexp, target, re.IGNORECASE):
    continue
  date = row[0].value
  amount = row[2].value
  category, account = get_defaults(target)
  data.append({"target": target, "date": date, "amount": amount, "category": category, "account": account})

if data:
  strings = map(lambda d: f"""('{d["target"]}', '{d["date"]}', {d['amount']}, '{d["category"]}', '{d["account"]}', '{iban}', '{currency}')""", data)
  values = ",\n".join(strings)
  print(f"INSERT OR IGNORE INTO ledger(target, date, amount, category, account, iban, currency) VALUES\n{values};")
