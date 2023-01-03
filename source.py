from __future__ import annotations
from dataclasses import dataclass
from datetime import datetime
from collections.abc import Sequence
from itertools import chain
from typing import List
import re
import util

@dataclass
class Row:
  target: str
  date: datetime
  amount: float
  iban: str
  currency: str

@dataclass
class AugmentedRow:
  target: str
  date: datetime
  amount: float
  iban: str
  currency: str
  category: str
  account: str

class DataSource:
  def create(bankname, *args):
    cls = {
      "Credo": CredoSource,
      "TBC": TBCSource
    }[bankname]
    return cls(*args)

  def get_rows(self) -> Sequence[Row]:
    pass

  def __init__(self, ignored_ibans: Sequence[str]):
    # Skip it if you don't plan to use _skip_target
    self.skip_targets_regexp = util.seq_to_regexp(self._get_specific_skip_targets() + ignored_ibans)

  def _get_specific_skip_targets(self) -> Sequence[str]:
    return []

  def _skip_target(self, target):
    return bool(re.search(self.skip_targets_regexp, target, re.IGNORECASE))

class CredoSource(DataSource):
  def __init__(self, filename, ignored_ibans = []):
    DataSource.__init__(self, ignored_ibans)
    import openpyxl
    account_sheet, self.transactions_sheet = openpyxl.load_workbook(filename = filename).worksheets
    self.iban = account_sheet.cell(row=3, column=2).value
    self.currency = account_sheet.cell(row=4, column=2).value

  def _get_specific_skip_targets(self):
    return ["Конвертация суммы", "Convert amount", "ანაბრის გახსნა", "ავტომატური შევსების ოპერაცია", "უნაღდო კონვერტაცია", "დეპოზიტის თანხის გატანა"]

  def get_rows(self):
    result = []
    iter = self.transactions_sheet.iter_rows()
    next(iter) # skip title

    for row in iter:
      if not row[2].value:
        continue # skip incomes

      target = self._get_target(row)

      if self._skip_target(target):
        continue

      date = row[0].value
      amount = row[2].value
      result.append(Row(target, date, amount, self.iban, self.currency))

    return result

  def _get_target(self, row):
    target = row[5].value
    target = re.sub("^გადახდა - ", "", target) # "Payment - "
    target = re.sub(" \d+\.\d{2} GEL \d{2}.\d{2}.\d{4}$", "", target)
    if target == "Personal Transfer.":
      beneficiary_name = row[6].value
      beneficiary_acc_num = row[7].value
      target = f"Personal transfer to {beneficiary_name} [{beneficiary_acc_num}]"
    return target

class TBCSource(DataSource):
  def __init__(self, filename, ignored_ibans = []):
    DataSource.__init__(self, ignored_ibans)
    import camelot
    self.table_list = camelot.read_pdf(filename, pages="1-end")
    self.iban = self._get_iban(filename)
    self.currency = self.table_list[0].df[1][0].split("\n")[1]

  def _get_specific_skip_targets(self):
    return ["Currency Exchange", "კონვერტაცია", "საკომისიო გადარიცხვებზე სხვა ბანკებში"]

  def _get_iban(self, filename):
    import PyPDF2
    with open(filename, "rb") as file:
      reader = PyPDF2.PdfFileReader(file)
      text = reader.getPage(0).extractText()
    return re.search("Account  Statement:\nანგარიშის მფლობელი:(.*) ", text)[1]

  def get_rows(self):
    result = []
    iter = chain(*(s.df.iterrows() for s in self.table_list[2:]))

    for _, row in iter:
      if row[0] == "თარიღი\nDate":
        continue # skip title
      if not row[2]:
        continue # skip incomes

      match = re.search("POS - ([^,]+),\s*თანხა (?:[\d.]+ \w+), ([^,]+)", row[1])
      if match:
        target, date = match.groups()
        date = datetime.strptime(date, "%b %d %Y %I:%M%p")
      else:
        target = row[1]
        date = datetime.strptime(row[0], "%d/%m/%Y")
      target = target.replace("\n", " ")

      if self._skip_target(target):
        continue

      amount = float(row[2].replace(",", ""))
      result.append(Row(target, date, amount, self.iban, self.currency))

    return result
